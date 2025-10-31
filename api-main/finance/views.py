from sqlite3 import IntegrityError
from accounts.enums import StudentShiftStatus
from accounts.models import StudentClassShift, StudentRequestShift
from organization.models import ClassRoomShift
from finance.enums import PaymentPurpose, PaymentStatus
from commons.utils.chapa_payment import retryPayment
from commons.utils.permissions import AllowGetOrCustomPermission, AllowPostOrCustomPermission
from commons.utils.filter_utils import fields_lookups
from rest_framework.filters import OrderingFilter, SearchFilter
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import generics, status
from commons.utils.filter_utils import fields_lookups
from commons.utils.permissions import AllowGetOrCustomPermission, AllowPostOrCustomPermission
from rest_framework.filters import OrderingFilter, SearchFilter
from rest_framework.response import Response
from rest_framework.exceptions import ValidationError
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import generics, status
from django.shortcuts import get_object_or_404, render
from finance.serializers import PaymentSerializer
from finance.models import Payment
from finance.serializers import PaymentSerializer
from finance.models import Payment
from finance.serializers import PaymentWebhookEventSerializer
from finance.models import PaymentWebhookEvent
from finance.serializers import FeePackageSerializer
from finance.models import FeePackage
from django.conf import settings
import hmac
import hashlib
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .models import ChapaWebhookEvent, InvoiceCounter
from rest_framework.permissions import AllowAny
from rest_framework.permissions import IsAuthenticated
from django.http import FileResponse
from commons.utils.pdf_generator import generate_receipt_pdf
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.db import transaction
from num2words import num2words
from commons.utils.email_utils import send_email_in_thread, send_password_reset_link
from django.utils import timezone
from django.db.models import Sum, Count
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.utils.timezone import localtime
from datetime import datetime, date
from django.db.models.functions import TruncDate, TruncMonth, TruncYear
import logging

logger = logging.getLogger(__name__)


class PaymentListCreateView(generics.ListCreateAPIView):
    permission_classes = [AllowGetOrCustomPermission]
    serializer_class = PaymentSerializer
    queryset = Payment.objects.all()
    filter_backends = [DjangoFilterBackend, OrderingFilter, SearchFilter]
    search_fields = ['ref_number']
    ordering_fields = '__all__'
    filterset_fields = fields_lookups(Payment)

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        group_by_raw = request.GET.getlist("group_by")  # ['date,month']

        group_by_list = []
        for item in group_by_raw:
            group_by_list.extend(item.split(","))

        group_by_list = [g.strip() for g in group_by_list if g]

        grouped = bool(group_by_list)

        if request.GET.get("export") == "excel":
            return self.export_excel(queryset, group_by_list)

        if grouped:
            # Get grouped queryset and columns
            grouped_queryset, columns = self.get_grouped_queryset(queryset, group_by_list)

            # Paginate the grouped queryset
            page = self.paginate_queryset(grouped_queryset)
            paginated_data = list(page) if page is not None else list(grouped_queryset)

            # Totals on full filtered queryset
            totals = queryset.aggregate(
                sum_amount=Sum("amount"),
                sum_tot=Sum("tot_amount"),
                sum_vat=Sum("vat_amount"),
                sum_total_amount=Sum("total_amount"),
                count=Count("*")
            )

            response_data = {
                "results": paginated_data,
                "totals": totals,
                "columns": columns,
            }

            return self.get_paginated_response(response_data) if page is not None else Response(response_data)

        else:
            # Normal (ungrouped) queryset
            page = self.paginate_queryset(queryset)
            serializer = self.get_serializer(page if page is not None else queryset, many=True)

            totals = queryset.aggregate(
                amount=Sum("amount"),
                tot_amount=Sum("tot_amount"),
                vat_amount=Sum("vat_amount"),
                total_amount=Sum("total_amount"),
            )

            response_data = {
                "results": serializer.data,
                "totals": totals,
            }

            return self.get_paginated_response(response_data) if page is not None else Response(response_data)
    
    def get_grouped_queryset(self, queryset, group_by_list):
        """
        group_by_list: list of strings, e.g., ["month"] or ["status", "payment_for"]
        Returns: aggregated queryset (dicts) and column names
        """
        original_group_by = group_by_list[:] 
        date_fields = {"date": TruncDate, "month": TruncMonth, "year": TruncYear}
        annotated_groups = []
        for idx, g in enumerate(group_by_list):
            if g in date_fields:
                alias = f"group_{idx}"
                queryset = queryset.annotate(**{alias: date_fields[g]("payment_date")})
                annotated_groups.append(alias)
            else:
                annotated_groups.append(g)

        if annotated_groups:
            for idx, g in enumerate(original_group_by):
                if g == "fee_package":
                    annotated_groups[idx] = "fee_package__name"
            queryset = queryset.values(*annotated_groups).annotate(
                sum_amount=Sum("amount"),
                sum_vat=Sum("vat_amount"),
                sum_tot=Sum("tot_amount"),
                sum_total_amount=Sum("total_amount"),
                count=Count("id")
            )
            # Apply ordering from request.GET["ordering"], fallback to group fields
            ordering = self.request.GET.get("ordering")
            if ordering:
                ordering_fields = [f.strip() for f in ordering.split(",")]
                queryset = queryset.order_by(*ordering_fields)
            else:
                queryset = queryset.order_by(*annotated_groups)

            columns = []

            # Add dynamic group_by columns
            for idx, g in enumerate(annotated_groups):
                original = original_group_by[idx]
                if original in date_fields:
                    title = original.capitalize()
                else:
                    title = original.replace("_", " ").capitalize()

                columns.append({"title": title, "key": g})

            # Add summary columns
            columns += [
                {"title": "Total Amount", "key": "sum_amount"},
                {"title": "TOT", "key": "sum_tot"},
                {"title": "VAT", "key": "sum_vat"},
                {"title": "Grand Total", "key": "sum_total_amount"},
                {"title": "Count", "key": "count"},
            ]
        else:
            columns = [
                {"title": "Ref Number", "key": "ref_number"},
                {"title": "User", "key": "user__username"},
                {"title": "Amount", "key": "amount"},
                {"title": "TOT", "key": "tot_amount"},
                {"title": "VAT", "key": "vat_amount"},
                {"title": "Total Amount", "key": "total_amount"},
                {"title": "Payment Date", "key": "payment_date"},
            ]
        return queryset, columns

    def export_excel(self, queryset, group_by_list):
        grouped = bool(group_by_list)
        
        if grouped:
            grouped_queryset, columns = self.get_grouped_queryset(queryset, group_by_list)
            data_to_export = grouped_queryset 
            totals = queryset.aggregate(
                sum_amount=Sum("amount"),
                sum_tot=Sum("tot_amount"),
                sum_vat=Sum("vat_amount"),
                sum_total_amount=Sum("total_amount"),
                sum_count=Count("*")
            )
        else:
            columns = [
                {"title": "Ref Number", "key": "ref_number"},
                {"title": "User", "key": "user__full_name"},
                {"title": "Amount", "key": "amount"},
                {"title": "TOT", "key": "tot_amount"},
                {"title": "VAT", "key": "vat_amount"},
                {"title": "Total Amount", "key": "total_amount"},
                {"title": "Payment Date", "key": "payment_date"},
            ]
            data_to_export = queryset.values(
                "ref_number", 
                "user__username", 
                "amount", 
                "tot_amount", 
                "vat_amount", 
                "total_amount", 
                "payment_date"
            )
            totals = queryset.aggregate(
                sum_amount=Sum("amount"),
                sum_tot=Sum("tot_amount"),
                sum_vat=Sum("vat_amount"),
                sum_total_amount=Sum("total_amount"),
                sum_count=Count("*")
            )

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Payments Report"

        # Header row
        for col_num, col in enumerate(columns, 1):
            ws.cell(row=1, column=col_num, value=col["title"])

        # Data rows
        for row_num, row in enumerate(data_to_export, 2):
            for col_num, col in enumerate(columns, 1):
                key = col["key"]
                value = row.get(key, "") if isinstance(row, dict) else getattr(row, key, "")
                if isinstance(value, (datetime, date)):
                    if(col['title'] == 'Month'):
                        value = value.strftime("%Y-%m")
                    elif(col['title'] == 'Year'):
                        value = value.strftime("%Y")
                    else:
                        value = value.strftime("%Y-%m-%d")
                elif key == "status":
                    value = PaymentStatus(value).name
                elif key == "payment_for":
                    value = PaymentPurpose(value).name
                elif key == "fee_package":
                    value = value.name
                ws.cell(row=row_num, column=col_num, value=value)

        # Totals row
        totals_row = ws.max_row + 1
        ws.cell(row=totals_row, column=1, value="Totals")
        for col_num, col in enumerate(columns, 1):
            key = col["key"]
            if key == "sum_amount":
                ws.cell(row=totals_row, column=col_num, value=totals["sum_amount"] or 0)
            elif key == "sum_tot":
                ws.cell(row=totals_row, column=col_num, value=totals["sum_tot"] or 0)
            elif key == "sum_vat":
                ws.cell(row=totals_row, column=col_num, value=totals["sum_vat"] or 0)
            elif key == "sum_total_amount":
                ws.cell(row=totals_row, column=col_num, value=totals["sum_total_amount"] or 0)
            elif key == "count":
                ws.cell(row=totals_row, column=col_num, value=totals["sum_count"] or 0)

        # Adjust column widths
        for i, _ in enumerate(columns, 1):
            ws.column_dimensions[get_column_letter(i)].width = 18

        # Prepare response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="PaymentsReport.xlsx"'
        wb.save(response)
        return response

class PaymentDetailView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [AllowGetOrCustomPermission]
    queryset = Payment.objects.all()
    serializer_class = PaymentSerializer

class PaymentWebhookEventListCreateView(generics.ListCreateAPIView):
    permission_classes = [AllowGetOrCustomPermission]
    serializer_class = PaymentWebhookEventSerializer
    queryset = PaymentWebhookEvent.objects.all()
    filter_backends = [DjangoFilterBackend, OrderingFilter, SearchFilter]
    search_fields = ['event_type', 'status']
    ordering_fields = '__all__'
    filterset_fields = fields_lookups(PaymentWebhookEvent)

class PaymentWebhookEventDetailView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [AllowGetOrCustomPermission]
    queryset = PaymentWebhookEvent.objects.all()
    serializer_class = PaymentWebhookEventSerializer

class FeePackageListCreateView(generics.ListCreateAPIView):
    permission_classes = [AllowGetOrCustomPermission]
    serializer_class = FeePackageSerializer
    queryset = FeePackage.objects.all()
    filter_backends = [DjangoFilterBackend, OrderingFilter, SearchFilter]
    search_fields = ['name', 'description']
    ordering_fields = '__all__'
    filterset_fields = fields_lookups(FeePackage)

class FeePackageDetailView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [AllowGetOrCustomPermission]
    queryset = FeePackage.objects.all()
    serializer_class = FeePackageSerializer


class ChapaWebhookRetryView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, payment_id):
        try:
            payment = Payment.objects.get(id=payment_id, user=request.user)
        except Payment.DoesNotExist:
            return Response({"error": "Payment not found."}, status=404)

        result = retryPayment(payment.id)

        if "error" in result:
            return Response({"error": result["error"]}, status=400)

        checkout_url = result["payment"].checkout_url
        if not checkout_url:
            return Response({"error": "Checkout URL not available."}, status=400)

        return Response({"checkout_url": checkout_url})

class ChapaWebhookView(APIView):
    permission_classes = [AllowAny]
    @transaction.atomic
    def post(self, request):
        secret = getattr(settings, 'CHAPA_WEBHOOK_SECRET', None)
        sig = request.headers.get('Chapa-Signature') or request.headers.get('x-chapa-signature')
        if secret and sig:
            digest = hmac.new(secret.encode(), request.body, hashlib.sha256).hexdigest()
            if not hmac.compare_digest(digest, sig):
                return Response({'detail': 'Invalid signature'}, status=status.HTTP_403_FORBIDDEN)

        payload = request.data
        event = payload.get('event')
        event_type = payload.get('type')

        obj = ChapaWebhookEvent.objects.create(
            event=event,
            event_type=event_type,
            account_name=payload.get('account_name'),
            account_number=payload.get('account_number'),
            bank_id=str(payload.get('bank_id')) if payload.get('bank_id') is not None else None,
            bank_name=payload.get('bank_name'),
            amount=payload.get('amount'),
            charge=payload.get('charge'),
            currency=payload.get('currency'),
            status=payload.get('status'),
            reference=payload.get('reference'),
            chapa_reference=payload.get('chapa_reference'),
            bank_reference=payload.get('bank_reference'),
            created_at=payload.get('created_at'),
            updated_at=payload.get('updated_at'),
            first_name=payload.get('first_name'),
            last_name=payload.get('last_name'),
            email=payload.get('email'),
            mobile=payload.get('mobile'),
            mode=payload.get('mode'),
            tx_ref=payload.get('tx_ref'),
            payment_method=payload.get('payment_method'),
            customization=payload.get('customization'),
            meta=str(payload.get('meta')) if payload.get('meta') is not None else None,
            row_data=payload
        )

        tx_ref = payload.get('tx_ref')
        webhook_status = payload.get('event', '').lower()

        try:
            payment = Payment.objects.get(ref_number=tx_ref)
        except Payment.DoesNotExist:
            payment = None
        if payment:
            # Map webhook status to your PaymentStatus enum
            if webhook_status == "charge.success":
                payment.status = PaymentStatus.PAID

                if not payment.invoice_no:
                    counter, _ = InvoiceCounter.objects.select_for_update().get_or_create(
                        id=1, defaults={"counter": 1}
                    )
                    payment.invoice_no = f"{counter.counter:08d}"
                    counter.counter += 1
                    counter.save()
                    payment.save()

                    if getattr(payment, "payment_for", None) == "TUITION":
                        for detail in payment.shift_details.select_related("student_shift", "fee_package"):
                            student_shift_request = detail.student_shift
                            student = student_shift_request.student

                            request_shifts = StudentRequestShift.objects.filter(student=student)
                            matching_shifts = ClassRoomShift.objects.none()

                            for rs in request_shifts:
                                qs = ClassRoomShift.objects.filter(
                                    class_room__institution=rs.institution,
                                    shift_days=rs.shift_days,
                                    start_time=rs.start_time,
                                    end_time=rs.end_time,
                                )
                                matching_shifts = matching_shifts | qs

                            matching_shifts = matching_shifts.distinct()

                            # Exclude full ones
                            occupied_ids = StudentClassShift.objects.filter(
                                class_room_shift__in=matching_shifts,
                                status=StudentShiftStatus.ACTIVE
                            ).values_list("class_room_shift_id", flat=True)

                            available_shifts = matching_shifts.exclude(id__in=occupied_ids)

                            # 2. Pick the first available shift
                            available_shift = available_shifts.first()
                            if not available_shift:
                                logger.warning(
                                    f"No available shifts for student {student.id} (payment {payment.id})"
                                )
                                continue

                            exists = StudentClassShift.objects.filter(
                                student=student, class_room_shift=available_shift
                            ).exists()
                            if not exists:
                                try:
                                    StudentClassShift.objects.create(
                                        student=student,
                                        institution=student.institution,
                                        class_room=available_shift.class_room,
                                        class_room_shift=available_shift,
                                        student_request_class_shift=student_shift_request,
                                        status=StudentShiftStatus.ACTIVE,
                                    )
                                except IntegrityError:
                                    pass


                # Build transactions list
                transactions = []
                if payment.shift_details.exists():
                    # Multiple shift payments
                    for detail in payment.shift_details.select_related('fee_package'):
                        transactions.append({
                            "no": detail.fee_package.item_code if detail.fee_package else "N/A",
                            "product": detail.fee_package.name if detail.fee_package else "Unknown",
                            "amount": float(detail.amount),
                        })
                else:
                    # Single payment (e.g., application, teacher)
                    transactions.append({
                        "no": payment.fee_package.item_code if getattr(payment, 'fee_package', None) else "N/A",
                        "product": payment.fee_package.name if getattr(payment, 'fee_package', None) else "Payment",
                        "amount": float(payment.total_amount),
                    })

                data = {
                    "invoice_no": str(obj.reference),
                    "local_invoice_no": str(payment.invoice_no),
                    "date": payment.create_date,
                    "payer_name": obj.account_name or f"{obj.first_name} {obj.last_name}" or "Unknown",
                    "payer_account": obj.account_number or obj.mobile or "N/A",
                    "payer_tin": payment.tin_no or "N/A",
                    "payer_address": payment.address or "N/A",
                    "payer_phone": getattr(obj, "mobile", "N/A"),
                    "vat_number": payment.vat_reg_no or "N/A",
                    "vat_date": payment.vat_reg_date or "N/A",
                    "transactions": transactions,
                    "service_fee": float(payment.service_fee),
                    "tot": float(payment.tot_amount),
                    "vat": float(payment.vat_amount),
                    "total": float(payment.total_amount),
                    "amount_in_words": (num2words(payment.total_amount) or "Amount in words here") + " ETB",
                    "payment_method": obj.payment_method or "Account",
                    "payment_status": obj.status or "Success",
                }

                # Generate PDF receipt
                pdf_file = generate_receipt_pdf(data)
                file_name = f"receipt_{payment.ref_number}.pdf"
                default_storage.save(file_name, ContentFile(pdf_file.getvalue()))

                # Save receipt path and payment date
                payment.receipt = file_name
                payment.payment_date = timezone.now()
                payment.save()
                user = payment.user
                subject = "Your Payment Receipt - Take The Stage"
                body_text = "Your payment was successful. Please find your receipt attached."
                body_html = f"""
                    <html>
                    <body>
                        <h2>Hi {user.first_name},</h2>
                        <p>Thank you for your payment.</p>
                        <p>Your receipt is attached to this email.</p>
                        <p><strong>Invoice No:</strong> {payment.invoice_no}</p>
                        <p><strong>Amount Paid:</strong> {payment.total_amount}</p>
                        <br/>
                        <p>Take The Stage</p>
                    </body>
                    </html>
                """
                send_email_in_thread(user, subject, body_text, body_html, pdf_file)
                send_password_reset_link(user.email)

            elif webhook_status in ["failed", "error"]:
                payment.status = PaymentStatus.REJECTED
                payment.save()
            elif webhook_status in ["pending", "processing"]:
                payment.status = PaymentStatus.PENDING
                payment.save()

        return Response({'status': 'received', 'id': obj.id}, status=status.HTTP_200_OK)

# class ReceiptPrintAPIView(APIView):
#     permission_classes = [AllowAny]
#     def get(self, request, payment_id):
#         payment = get_object_or_404(Payment, id=payment_id)
#         if not payment.receipt or not default_storage.exists(payment.receipt):
#             return Response({"detail": "Receipt not found."}, status=404)
#         return FileResponse(default_storage.open(payment.receipt), as_attachment=True, filename=f"receipt_{payment.ref_number}.pdf")